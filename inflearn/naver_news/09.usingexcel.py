from openpyxl import Workbook

# 엑셀 생성하기
wb = Workbook()

# 엑셀 시트 생성하기
ws = wb.create_sheet("testsheet")

# 엑셀 데이터 저장하기
ws['A1'] = "test"

# 엑셀 저장하기
wb.save("test.xlsx")

