#패키기 가져오기
import openpyxl
import requests
from bs4 import BeautifulSoup

#workbook 생성
wb = openpyxl.Workbook()
# Sheet 활성화
sheet = wb.active
sheet.cell(row = 1, column = 1).value = "1.1"
sheet.cell(row=1,column=2).value ="1.2"
sheet.cell(row=1,column=3).value ="1.3"
# 데이터 저장
wb.save("testfile.xlsx")